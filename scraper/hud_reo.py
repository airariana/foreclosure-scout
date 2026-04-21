"""
HUD REO (Real Estate Owned) source for Foreclosure Scout.

HUD HomeStore's Excel export flow is the data source. Because the export
button requires a logged-in HUD account session, we can't automate the
download — instead the user manually exports per state and commits the
files at `data/hud_{state}.xlsx`. This module reads whatever's there on
each scheduled run, one file per state.

HUD data quality is significantly better than trustee-firm scrapes:
  - Actual list price (not estimated) → HIGH-confidence pricing
  - Real beds / baths / sqft / year built — no enrichment needed
  - FHA financing indicator ("IN (Insured)", "IE (Insured Escrow)", "UI (Uninsured)")
  - List date + bid deadline

Expected spreadsheet columns:
  Property Case | Address | City | State | Zip Code | County |
  Price | Bed | Bath | Square Footage | Year Built | FHA Financing |
  List Date | Bid Open Date | Listing Period | Status | Period Deadline
"""

from __future__ import annotations

import hashlib
import logging
import re
from datetime import datetime, date
from pathlib import Path

log = logging.getLogger(__name__)

# Per-state xlsx exports. Add a new state by dropping a matching file here
# and appending to this list — the loader reads each one independently.
HUD_XLSX_FILES = [
    (Path("data/hud_va.xlsx"), "VA"),
    (Path("data/hud_md.xlsx"), "MD"),
    (Path("data/hud_dc.xlsx"), "DC"),
]

ALLOWED_HUD_STATES = {"VA", "MD", "DC"}


def scrape_hud_reo() -> list[dict]:
    """Parse all per-state HUD REO xlsx exports into v2.1-schema property dicts."""
    log.info("Reading HUD REO from xlsx ...")
    properties: list[dict] = []

    try:
        from openpyxl import load_workbook
    except ImportError:
        log.warning("HUD REO: openpyxl not installed — skipping")
        return properties

    for path, state_hint in HUD_XLSX_FILES:
        if not path.exists():
            log.info(f"HUD REO: no file at {path} — skipping {state_hint}")
            continue

        try:
            added_before = len(properties)
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                log.warning(f"HUD REO {state_hint}: xlsx is empty")
                continue

            # Build header → column-index map so reordering doesn't break us.
            header = [str(h or "").strip() for h in rows[0]]
            idx = {name: i for i, name in enumerate(header)}

            required = {"Address", "City", "State", "Zip Code", "Price"}
            missing = required - set(idx.keys())
            if missing:
                log.error(f"HUD REO {state_hint}: missing expected columns: {missing}")
                continue

            for raw_row in rows[1:]:
                if not raw_row or not raw_row[idx["Address"]]:
                    continue
                state = str(raw_row[idx.get("State", -1)] or "").strip().upper()
                if state not in ALLOWED_HUD_STATES:
                    continue

                prop = _build_property(raw_row, idx)
                if prop:
                    properties.append(prop)

            log.info(f"HUD REO {state_hint}: {len(properties) - added_before} properties from {path.name}")

        except Exception as e:
            log.error(f"HUD REO {state_hint} parse failed: {e}")
            continue

    log.info(f"HUD REO: {len(properties)} total properties across VA/MD/DC")
    return properties


def _build_property(row: tuple, idx: dict) -> dict | None:
    """Convert one xlsx row into the shared property schema."""
    def g(col: str):
        if col not in idx:
            return None
        v = row[idx[col]]
        return v if v not in (None, "") else None

    def g_int(col: str) -> int | None:
        v = g(col)
        if v is None:
            return None
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return None

    def g_str(col: str) -> str:
        v = g(col)
        return str(v).strip() if v is not None else ""

    address   = g_str("Address")
    city      = g_str("City")
    state     = g_str("State").upper() or "VA"
    zip_code  = g_str("Zip Code")
    county_raw = g_str("County")
    case_no   = g_str("Property Case")

    if not address:
        return None

    # HUD's Price is the actual list/bid price — this unlocks HIGH-confidence.
    list_price = g_int("Price")
    beds       = g_int("Bed")
    baths_raw  = g("Bath")
    try:
        baths = float(baths_raw) if baths_raw is not None else 0
    except (TypeError, ValueError):
        baths = 0
    sqft       = g_int("Square Footage")
    year_built = g_int("Year Built")
    fha_status = g_str("FHA Financing")
    hud_status = g_str("Status")

    # Parse dates — HUD format is M/D/YYYY.
    def parse_date(s: str) -> str | None:
        if not s:
            return None
        for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except Exception:
                continue
        return None

    list_date     = parse_date(g_str("List Date"))
    bid_open_date = parse_date(g_str("Bid Open Date"))
    deadline_date = parse_date(g_str("Period Deadline"))

    # Use bid-open as the "sale date" for consistency with trustee sources
    # (it's the primary date the investor cares about).
    sale_date = bid_open_date or list_date
    days_to_sale = _days_until(sale_date)

    # Normalize county → "X County" or "X City" shape, and route through the
    # shared normalizer if imported from the main scraper. We inline a light
    # version here to avoid a circular import.
    county = _normalize_county(county_raw)

    # HUD-specific ARV heuristic: HUD list price is typically 5-15% below
    # market (they want quick sale, not max price). So ARV ≈ list_price * 1.10
    # as a conservative estimate. This is better than the county-base derivation
    # we do for trustee sales because we have the REAL list price signal.
    arv_estimate = int(round(list_price * 1.10)) if list_price else 0

    # Investment metrics — will be refined by build_pricing() in the main
    # scraper, but provide defaults so downstream code has numbers.
    monthly_rent_estimate = int(round(arv_estimate * 0.007)) if arv_estimate else 0
    mortgage_monthly      = int(round(list_price * 0.006)) if list_price else 0
    prop_tax              = int(round(arv_estimate * 0.009 / 12)) if arv_estimate else 0
    insurance             = int(round(arv_estimate * 0.005 / 12)) if arv_estimate else 0
    vacancy               = int(round(monthly_rent_estimate * 0.08))
    cash_flow             = monthly_rent_estimate - mortgage_monthly - prop_tax - insurance - vacancy
    noi_annual            = (monthly_rent_estimate - prop_tax - insurance - vacancy) * 12
    cap_rate              = round((noi_annual / list_price * 100), 1) if list_price else 0

    # 70% rule + DSCR for HUD (same benchmarks as trustee sources)
    rehab_assumed = int(round(arv_estimate * 0.08))
    mao_70 = int(round(arv_estimate * 0.70 - rehab_assumed)) if arv_estimate else 0
    gap_to_mao = int((list_price or 0) - mao_70)
    passes_70 = bool(list_price) and list_price <= mao_70
    dscr = round(monthly_rent_estimate / mortgage_monthly, 2) if mortgage_monthly > 0 else 0

    score = _compute_score(list_price, arv_estimate, cash_flow, cap_rate, county)
    # Apply the same research-based bonuses/penalties as main scraper
    if passes_70:
        score += 5
    if dscr >= 1.25:
        score += 5
    elif dscr < 1.0 and dscr > 0:
        score -= 5
    if not sqft:
        score -= 5
    score = max(0, min(100, score))

    if score >= 90:   grade = "A+"
    elif score >= 80: grade = "A"
    elif score >= 70: grade = "B"
    elif score >= 60: grade = "C"
    else:             grade = "D"

    pricing = {
        "eav":                     list_price or 0,
        "arv":                     arv_estimate,
        "confidence":              "HIGH — HUD list price" if list_price else "LOW — no HUD price",
        "county_base":             None,
        "type_multiplier":         1.0,
        "opening_bid":             list_price,
        "original_loan":           None,
        "monthly_rent_estimate":   monthly_rent_estimate,
        "rent_source":             "heuristic",
        "cash_flow_estimate":      cash_flow,
        "cap_rate":                cap_rate,
        "discount_to_arv":         round((1 - list_price / arv_estimate) * 100, 1) if list_price and arv_estimate else 0,
        "mao_70":                  mao_70,
        "gap_to_mao":              gap_to_mao,
        "passes_70_rule":          passes_70,
        "dscr":                    dscr,
        "score":                   score,
        "grade":                   grade,
    }

    tags = [f"{state} Foreclosure", "HUD REO"]
    if fha_status:
        tags.append(f"FHA: {fha_status}")

    return {
        "id":               _make_id("hud", address, sale_date or ""),
        "source":           "HUD HomeStore",
        "source_url":       "https://www.hudhomestore.gov/",
        "firm_file_number": case_no or None,
        "address":          address,
        "city":             city,
        "state":            state,
        "zip_code":         zip_code,
        "county":           county,
        "lat":              None,
        "lng":              None,
        "sale_date":        sale_date,
        "sale_date_raw":    g_str("Bid Open Date") or g_str("List Date"),
        "sale_time":        None,
        "sale_location":    "HUD online bid submission",
        "listingType":      "HUD Home",  # HUD HomeStore = government-owned REO (blue pin, distinct from private bank REO)
        "property_type":    "Single Family",  # HUD REO is overwhelmingly SFH
        "sqft":             sqft,
        "beds":             beds,
        "baths":            baths,
        "year_built":       year_built,
        "pricing":          pricing,
        "tags":             tags,
        "status":           hud_status or "active",
        "hud_list_date":    list_date,
        "hud_bid_open":    bid_open_date,
        "hud_deadline":    deadline_date,
        "hud_fha":         fha_status,
        "scraped_at":       datetime.utcnow().isoformat() + "Z",
        "days_to_sale":     days_to_sale,
    }


# ── Helpers ──────────────────────────────────────────────────────────────────

def _days_until(iso_date: str | None) -> int | None:
    if not iso_date:
        return None
    try:
        target = datetime.strptime(iso_date, "%Y-%m-%d").date()
        return (target - date.today()).days
    except Exception:
        return None


def _normalize_county(raw: str) -> str:
    raw = raw.strip()
    if not raw:
        return "Unknown County"
    known_cities = {
        "Alexandria", "Fairfax", "Falls Church", "Manassas", "Manassas Park",
        "Richmond", "Norfolk", "Virginia Beach", "Chesapeake", "Newport News",
        "Hampton", "Portsmouth", "Suffolk", "Poquoson", "Fredericksburg",
        "Colonial Heights", "Petersburg", "Lynchburg", "Roanoke",
    }
    for city in known_cities:
        if city.lower() == raw.lower():
            return f"{city} City"
    if "County" in raw or "City" in raw:
        return raw
    return f"{raw} County"


def _make_id(prefix: str, address: str, sale_date: str) -> str:
    seed = f"{prefix}:{address.lower().strip()}:{sale_date}"
    return f"va-{prefix}-{hashlib.md5(seed.encode()).hexdigest()[:8]}"


# Tier-1 / Tier-2 counties for investment scoring — mirrors main scraper.
TIER_1 = {
    "Fairfax County", "Arlington County", "Loudoun County", "Prince William County",
    "Alexandria City", "Falls Church City", "Manassas City",
}
TIER_2 = {
    "Henrico County", "Chesterfield County", "Hanover County", "Richmond City",
    "Virginia Beach City", "Chesapeake City", "Norfolk City", "Newport News City",
    "Hampton City", "Stafford County", "Spotsylvania County", "Fredericksburg City",
}


def _compute_score(list_price: int, arv: int, cash_flow: int, cap_rate: float, county: str) -> int:
    """0-100 investment score mirroring main scraper's build_pricing logic."""
    if not list_price or not arv:
        return 0
    discount_to_arv = (1 - list_price / arv) * 100
    score = 0
    score += min(30, discount_to_arv * 1.2)
    score += min(25, max(0, cap_rate * 3))
    score += min(20, max(0, cash_flow / 50))
    if county in TIER_1:
        score += 15
    elif county in TIER_2:
        score += 10
    else:
        score += 5
    score += 10  # HUD list price = HIGH confidence bonus
    return min(100, round(score))
